import json
import os
import tempfile
import base64
from datetime import datetime, timedelta, timezone

def _get_mgn_now() -> datetime:
    """Возвращает текущее время в Магнитогорске (GMT+5) без информации о таймзоне для совместимости."""
    tz = timezone(timedelta(hours=5))
    return datetime.now(tz).replace(tzinfo=None)

DB_FILE = "/data/users.json" if os.path.isdir("/data") else "users.json"

# Внутрипамятый кэш для исключения постоянного чтения с диска
_users_cache = None


def _xor_cipher(data_str: str, key: str) -> str:
    key_bytes = key.encode("utf-8")
    data_bytes = data_str.encode("utf-8")
    xor_bytes = bytearray(d ^ key_bytes[i % len(key_bytes)] for i, d in enumerate(data_bytes))
    return base64.b64encode(xor_bytes).decode("utf-8")


def _xor_decipher(encoded_str: str, key: str) -> str:
    key_bytes = key.encode("utf-8")
    xor_bytes = base64.b64decode(encoded_str.encode("utf-8"))
    data_bytes = bytearray(d ^ key_bytes[i % len(key_bytes)] for i, d in enumerate(xor_bytes))
    return data_bytes.decode("utf-8")


def _load() -> dict:
    if not os.path.exists(DB_FILE):
        return {}
    
    from config import DB_ENCRYPTION_KEY
    
    with open(DB_FILE, "r", encoding="utf-8") as f:
        content = f.read().strip()
        if not content:
            return {}
            
        # Если начинается с '{', то это старый незашифрованный формат
        if content.startswith("{"):
            try:
                return json.loads(content)
            except Exception:
                return {}
        
        # Пытаемся расшифровать
        try:
            decrypted = _xor_decipher(content, DB_ENCRYPTION_KEY)
            return json.loads(decrypted)
        except Exception:
            return {}


def _save(data: dict):
    from config import DB_ENCRYPTION_KEY
    data_str = json.dumps(data, ensure_ascii=False, indent=2)
    encrypted = _xor_cipher(data_str, DB_ENCRYPTION_KEY)
    
    with open(DB_FILE, "w", encoding="utf-8") as f:
        f.write(encrypted)


def _get_cache() -> dict:
    global _users_cache
    if _users_cache is None:
        _users_cache = _load()
    return _users_cache


def get_user_group(user_id: int) -> str | None:
    """Возвращает сохранённую группу пользователя или None из кэша памяти."""
    cache = _get_cache()
    info = cache.get(str(user_id))
    if isinstance(info, dict):
        return info.get("group")
    return info  # string or None


def _sync_user_to_google(user_id: int, info: dict):
    """Отправляет данные пользователя в Google Таблицу в фоновом режиме."""
    from config import GOOGLE_SHEET_URL
    if not GOOGLE_SHEET_URL:
        return
        
    import asyncio
    import aiohttp
    
    async def task():
        payload = {
            "user_id": str(user_id),
            "group": info.get("group"),
            "username": info.get("username"),
            "first_name": info.get("first_name"),
            "last_name": info.get("last_name"),
            "joined_at": info.get("joined_at"),
            "last_seen": info.get("last_seen")
        }
        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(GOOGLE_SHEET_URL, json=payload, timeout=10) as resp:
                    await resp.read()
        except Exception:
            pass
            
    try:
        loop = asyncio.get_running_loop()
        loop.create_task(task())
    except RuntimeError:
        pass


def set_user_group(user_id: int, group: str):
    """Сохраняет группу пользователя в кэш и записывает на диск."""
    cache = _get_cache()
    uid_str = str(user_id)
    info = cache.get(uid_str)
    
    now_str = _get_mgn_now().strftime("%Y-%m-%d %H:%M:%S")
    
    if isinstance(info, dict):
        info["group"] = group
    else:
        info = {
            "group": group,
            "joined_at": now_str,
            "last_seen": now_str,
            "username": None,
            "first_name": None,
            "last_name": None
        }
    cache[uid_str] = info
    _save(cache)
    _sync_user_to_google(user_id, info)


def update_user_activity(user_id: int, username: str | None, first_name: str | None, last_name: str | None) -> bool:
    """Обновляет информацию об имени аккаунта и времени последней активности.
    Возвращает True, если пользователь абсолютно новый (ранее отсутствовал в базе)."""
    cache = _get_cache()
    uid_str = str(user_id)
    is_new = uid_str not in cache
    
    now_str = _get_mgn_now().strftime("%Y-%m-%d %H:%M:%S")
    
    info = cache.get(uid_str)
    if isinstance(info, dict):
        if "joined_at" not in info:
            info["joined_at"] = now_str
        info["last_seen"] = now_str
        info["username"] = username
        info["first_name"] = first_name
        info["last_name"] = last_name
    elif isinstance(info, str):
        info = {
            "group": info,
            "joined_at": now_str,
            "last_seen": now_str,
            "username": username,
            "first_name": first_name,
            "last_name": last_name
        }
    else:
        info = {
            "group": None,
            "joined_at": now_str,
            "last_seen": now_str,
            "username": username,
            "first_name": first_name,
            "last_name": last_name
        }
        
    cache[uid_str] = info
    _save(cache)
    _sync_user_to_google(user_id, info)
    return is_new


def get_admin_settings() -> dict:
    """Возвращает настройки администратора."""
    cache = _get_cache()
    settings = cache.get("__settings__")
    if not isinstance(settings, dict):
        settings = {"notify_new_users": True}
    return settings


def set_admin_settings(settings: dict):
    """Сохраняет настройки администратора в базу."""
    cache = _get_cache()
    cache["__settings__"] = settings
    _save(cache)


def _parse_datetime(dt_str: str | None) -> datetime | None:
    if not dt_str:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(dt_str, fmt)
        except ValueError:
            pass
    return None


def get_admin_stats() -> str:
    """Генерирует сводную текстовую статистику с ASCII-графиками."""
    cache = _get_cache()
    total_users = sum(1 for k in cache if k != "__settings__")
    
    now = _get_mgn_now()
    today_start = datetime(now.year, now.month, now.day)
    seven_days_ago = today_start - timedelta(days=7)
    thirty_days_ago = today_start - timedelta(days=30)
    
    today_count = 0
    week_count = 0
    month_count = 0
    
    for user_id, info in cache.items():
        if user_id == "__settings__":
            continue
        joined_str = None
        if isinstance(info, dict):
            joined_str = info.get("joined_at")
            
        dt = _parse_datetime(joined_str)
        if dt:
            if dt >= today_start:
                today_count += 1
            if dt >= seven_days_ago:
                week_count += 1
            if dt >= thirty_days_ago:
                month_count += 1
                
    def make_bar(val, max_val):
        if max_val == 0 or val == 0:
            return "░░░░░░░░░░"
        filled = min(10, int(val / max_val * 10))
        return "█" * filled + "░" * (10 - filled)
        
    max_val = max(today_count, week_count, month_count, 1)
    bar_today = make_bar(today_count, max_val)
    bar_week = make_bar(week_count, max_val)
    bar_month = make_bar(month_count, max_val)
    
    lines = [
        "📊 *Статистика пользователей бота:*",
        f"👥 Всего пользователей: *{total_users}*",
        "",
        "📈 *Прирост новых пользователей:*",
        f"📅 За сегодня:  `[{bar_today}]` *{today_count}* чел.",
        f"📅 За 7 дней:   `[{bar_week}]` *{week_count}* чел.",
        f"📅 За 30 дней:  `[{bar_month}]` *{month_count}* чел.",
    ]
    return "\n".join(lines)


def generate_users_report() -> str:
    """Генерирует красивый Excel-отчет о пользователях."""
    cache = _get_cache()
    total_users = sum(1 for k in cache if k != "__settings__")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    # Показывать сетку
    ws.views.sheetView[0].showGridLines = True
    
    # Заголовки таблицы
    headers = [
        "Telegram ID", 
        "Группа", 
        "Telegram Username", 
        "Имя и Фамилия", 
        "Зарегистрирован", 
        "Последняя активность"
    ]
    
    # Стили
    font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Темно-синий
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # Светлый серо-синий
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    border_double = Side(border_style="double", color="1F497D")
    
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    header_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_double)
    
    # Заголовок отчета
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Отчет по пользователям бота МАГПК Расписание"
    title_cell.font = font_title
    title_cell.alignment = align_left
    ws.row_dimensions[1].height = 40
    
    # Инфострока
    ws.merge_cells("A2:F2")
    info_cell = ws["A2"]
    info_cell.value = f"Всего пользователей: {total_users}  |  Дата генерации: {_get_mgn_now().strftime('%d.%m.%Y %H:%M:%S')}"
    info_cell.font = Font(name="Calibri", size=11, italic=True)
    info_cell.alignment = align_left
    ws.row_dimensions[2].height = 20
    
    # Заголовки (строка 4)
    ws.row_dimensions[4].height = 28
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = header_border
        
    # Данные (с 5 строки)
    row_num = 5
    for uid_str, info in cache.items():
        if uid_str == "__settings__":
            continue
        group = ""
        username = ""
        name = "Пользователь"
        joined = "-"
        last_seen = "-"
        
        if isinstance(info, dict):
            group = info.get("group") or ""
            username = f"@{info.get('username')}" if info.get("username") else "-"
            first = info.get("first_name") or ""
            last = info.get("last_name") or ""
            if first or last:
                name = f"{first} {last}".strip()
            joined = info.get("joined_at") or "-"
            last_seen = info.get("last_seen") or "-"
        else:
            group = info or ""
            
        row_data = [uid_str, group, username, name, joined, last_seen]
        
        ws.row_dimensions[row_num].height = 20
        is_even = (row_num % 2 == 0)
        
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.font = font_data
            cell.border = cell_border
            
            # Зебра-эффект
            if is_even:
                cell.fill = fill_zebra
                
            # Выравнивание
            if col_num in [1, 2, 5, 6]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
        row_num += 1
        
    # Автоподбор ширины колонок
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    filepath = os.path.join(tempfile.gettempdir(), "users_report.xlsx")
    wb.save(filepath)
    return filepath
